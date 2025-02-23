import pdfplumber
import re
from openpyxl import Workbook
from openpyxl.styles import numbers, Font
from getpdf import pdf

pdf_paths, excel_outputs = pdf()


def parent(pdf_path, excel_output):
    # Process data from a single page (extract and pass)
    def process_single_page(page):
        page_data = page.extract_text().split("\n")

        return page_data

    #Extract data from each page, process and return separate lists for each page
    def process_pdf_pages(pdf_path):
        all_page_data = []  
        
        with pdfplumber.open(pdf_path) as pdf:
            # Iterate through all pages in the PDF
            for page in pdf.pages:
                lines = process_single_page(page)
                
                all_page_data.append(lines)
        
        return all_page_data


    # Regex Patterns
    date_pattern = re.compile(r"\b\d{2}/\d{2}/\d{4}\b")  #Matches "25/12/2022"
    statement_period_pattern = re.compile(r"Period:\s*(\d{2}/\d{2}/\d{4})\s*to\s*(\d{2}/\d{2}/\d{4})")
    account_number_pattern = re.compile(r"ACCOUNT\s#:\s([A-Za-z]+)\s-\s(\d+)")
    amount_match = re.compile(r"\d{1,3}(?:,\d{3})*\.\d{2}\s*-?")
    irrelevant_patterns = [r"^Page:\s*\d+", r"^Total:?", r"^PERIODIC STATEMENT:?"]  


    def clean_match(match):
        cleaned_match = float(match.group().replace(',', ''))  # Extract, clean, and convert

        return cleaned_match

    nest_list = process_pdf_pages(pdf_path)

    #Extract details (Account holder information, Statement period, Opening balance, Closing balance)
    def extract_statement_details(nested_list):
        account_holder = None
        statement_period = None
        opening_balance = None
        deposits_credits_balance = None
        cheques_debits_balance = None
        closing_balance = None
        
        
        for lines in nested_list:
            for i, line in enumerate(lines):
                # Extract Account Holder Name
                name_match = re.match(r"^([A-Z\s]+)\s+Date:", line)
                if name_match:
                    account_holder = name_match.group(1).strip()

                # Extract Statement Period
                period_match = statement_period_pattern.search(line)
                if period_match:
                    statement_period = (period_match.group(1), period_match.group(2))

                # Search for the account number in the line
                account_match = account_number_pattern.search(line)
                if account_match:
                    account_type = account_match.group(1)  # Extract the account type
                    account_number = account_match.group(2)  # Extract the account number 
                    # Mask all but the last 4 digits 
                    if len(account_number) > 4:
                        account_number = 'XXXX-XXXX-' + account_number[-4:]
                    
                # Match opening balance
                if "Beginning Balance" in line and i + 1 < len(lines):
                    match = amount_match.search(lines[i + 1])
                    if match:
                        cleaned_match = clean_match(match)
                        opening_balance = cleaned_match

                # Match deposits
                elif "Deposits & Other Credits" in line:
                    match = amount_match.search(line)
                    if match:
                        cleaned_match = clean_match(match)
                        deposits_credits_balance = cleaned_match

                # Match cheques & debits
                elif "Cheques & Other Debits" in line:
                    match = amount_match.search(line)
                    if match:
                        cleaned_match = clean_match(match)
                        cheques_debits_balance = cleaned_match

                # Match closing balance
                elif "Ending Balance" in line and i + 1 < len(lines):
                    match = amount_match.search(lines[i + 1])
                    if match:
                        cleaned_match = clean_match(match)
                        closing_balance = cleaned_match

        summary = {
            "Account Holder": account_holder,
            "Account Type": account_type,
            "Account Number": account_number,
            "Statement Period Start": statement_period[0],
            "Statement Period End": statement_period[1],
            "Opening Balance": opening_balance,
            "Closing Balance": closing_balance,
            "Deposits & Other Credits": deposits_credits_balance,
            "Cheques & Other Debits": cheques_debits_balance
        }

        return summary

    def is_irrelevant_line(line):
        return any(re.search(pattern, line, re.IGNORECASE) for pattern in irrelevant_patterns)

    # If the transaction type is "Outgoing" make the amount negative
    def amount_negative(current_transaction, cleaned_amounts, amounts):
        current_transaction["Amount"] = -abs(cleaned_amounts[-2] if len(amounts) >= 2 else amounts[-1])

    # Extract all bank statement transactions
    def extract_transactions(nested_list):
        transactions = []
        incomplete_transactions = []
        current_transaction = None
        inside_transactions = False  # Track to see if inside the transactions table
        
        
        for lines in nested_list:
            for line in lines:
                if "TRANSACTION INFORMATION" in line:
                    continue
                
                # Detect table header to start capturing transactions
                if "Date" in line and "Description" in line and "Amount" in line and "Balance" in line:
                    inside_transactions = True
                    continue

                if inside_transactions:
                    amounts = amount_match.findall(line)  # Extract amounts

                    if amounts:
                        transaction_type = "Outgoing" if any("-" in amt for amt in amounts) else "Incoming"
                        
                        # Remove trailing "-" from outgoing amounts
                        cleaned_amounts = [amt.rstrip("-") for amt in amounts]
                        cleaned_amounts = [float(num.replace(',', '')) for num in cleaned_amounts]
                    
                    clean_line = re.sub(r"\s*-\s*$", "", amount_match.sub("", line)).strip()  # Remove amounts from text

                    # Identify a new transaction start
                    date_match = re.match(r"\d{2}/\d{2}", line)
                    if date_match and amounts:
                        if current_transaction and current_transaction["Amount"] is not None:
                            transactions.append(current_transaction)
                            if not current_transaction["Description"]:
                                incomplete_transactions.append(current_transaction)

                        date = date_match.group().strip()
                        description = clean_line.replace(date, "").strip()  # Remove date from description

                        # Start a new transaction
                        current_transaction = {
                            "Date": date,
                            "Description": description,
                            "Amount": cleaned_amounts[-2] if len(amounts) >= 2 else amounts[-1],
                            "Balance": cleaned_amounts[-1] if amounts else None,
                            "Type": transaction_type
                        }
                        
                        if transaction_type == "Outgoing":
                            current_transaction["Amount"] = -abs(cleaned_amounts[-2] if len(amounts) >= 2 else amounts[-1])
                    # If date but no amount description is wrapped
                    elif date_match and not amounts:
                        if current_transaction and current_transaction["Amount"] is not None:
                            transactions.append(current_transaction) 
                            if not current_transaction["Description"]:
                                incomplete_transactions.append(current_transaction)

                        date = date_match.group().strip()
                        description = clean_line.replace(date, "").strip()  # Remove date from description
                        
                        # Start a new transaction (no amount detected yet)
                        current_transaction = {
                            "Date": date,
                            "Description": description,
                            "Amount": None,
                            "Balance": None,
                            "Type": None
                        }
                    elif current_transaction:
                        if is_irrelevant_line(clean_line):
                            continue  # Skip irrelevant lines

                        # If the line contains amounts end the transaction
                        if amounts:
                            current_transaction["Description"] += " " + clean_line
                            current_transaction["Amount"] = cleaned_amounts[-2] if len(amounts) >= 2 else amounts[-1]
                            current_transaction["Balance"] = cleaned_amounts[-1] if amounts else None
                            current_transaction["Type"] = transaction_type
                            
                            if transaction_type == "Outgoing":
                                current_transaction["Amount"] = -abs(cleaned_amounts[-2] if len(amounts) >= 2 else amounts[-1])
                        else:
                            # If no amount is found it is a continuation of the description
                            current_transaction["Description"] += " " + clean_line

            inside_transactions = False  # Reset for the next page

        # Add the last transaction
        if current_transaction and current_transaction["Amount"] is not None:
            transactions.append(current_transaction)
            if not current_transaction["Description"]:
                incomplete_transactions.append(current_transaction)

        return transactions, incomplete_transactions

    def incoming_transactions(transactions):
        incoming = [t for t in transactions if t["Type"] == "Incoming"]

        return incoming

    def outgoing_transactions(transactions):
        outgoing = [t for t in transactions if t["Type"] == "Outgoing"]

        return outgoing

    def create_excel_sheet(summary, transactions, incomplete_transactions):

        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"

        # Adding the header
        headers = list(summary.keys())
        ws.append(headers)

        # Adding the summary values
        values = list(summary.values()) 
        ws.append(values)

        # Apply currency format 
        for row in ws.iter_rows(min_row=2, min_col=6, max_col=9):  
            for cell in row:
                if isinstance(cell.value, (int, float)):  
                    cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE  

        # Auto-adjust column widths
        for col in ws.iter_cols(min_col=1, max_col=ws.max_column):
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            ws.column_dimensions[column].width = max_length + 2

        # Create the transactions sheet
        ws_transactions = wb.create_sheet(title="Transactions")

        # Add transaction table headers
        ws_transactions.append(["Date", "Description", "Amount", "Balance"])

        # Add transactions
        for transaction in transactions:
            ws_transactions.append([transaction["Date"], transaction["Description"], transaction["Amount"], transaction["Balance"]])

        # Apply currency format
        for row in ws_transactions.iter_rows(min_row=2, min_col=3, max_col=4):  
            for cell in row:
                if isinstance(cell.value, (int, float)):  
                    cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE  

        # Automatically resize columns in transactions sheet
        for col in ws_transactions.iter_cols(min_col=1, max_col=ws_transactions.max_column):
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            ws_transactions.column_dimensions[column].width = max_length + 2

        # Create the transactions sheet
        ws_incomplete_transactions = wb.create_sheet(title="Incomplete Transactions")

        # Add transaction table headers
        ws_incomplete_transactions.append(["Date", "Description", "Amount", "Balance"])

        # Add transactions
        for transaction in incomplete_transactions:
            ws_incomplete_transactions.append([transaction["Date"], transaction["Description"], transaction["Amount"], transaction["Balance"]])

        # Apply currency format
        for row in ws_incomplete_transactions.iter_rows(min_row=2, min_col=3, max_col=4):  
            for cell in row:
                if isinstance(cell.value, (int, float)):  
                    cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE  

        # Automatically resize columns in transactions sheet
        for col in ws_incomplete_transactions.iter_cols(min_col=1, max_col=ws_incomplete_transactions.max_column):
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            ws_incomplete_transactions.column_dimensions[column].width = max_length + 2

        sheets = [ws, ws_transactions, ws_incomplete_transactions]
        
        # Apply bold formatting to the first row of sheets
        for sheet in sheets:  
            for cell in sheet[1]:
                cell.font = Font(bold=True)

        # Save workbook
        wb.save(excel_output)

        print("Success")

    transactions, incomplete_transactions = extract_transactions(nest_list)
    create_excel_sheet(extract_statement_details(nest_list), transactions, incomplete_transactions)

    return process_pdf_pages, extract_statement_details, extract_transactions, incoming_transactions, outgoing_transactions

# Iterate over both lists simultaneously
for pdf_file, excel_file in zip(pdf_paths, excel_outputs):
    parent(pdf_file, excel_file)